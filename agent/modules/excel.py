"""
Excel automation module — read, write, and manipulate .xlsx files.

Contract: exposes TOOL_DEFINITIONS and dispatch() for the registry.
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TOOL_DEFINITIONS: list[dict] = [
    {
        "name": "excel_read_sheet",
        "description": "Read rows from a sheet in an .xlsx file. Returns headers and rows as lists.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to .xlsx file."},
                "sheet": {"type": "string", "description": "Sheet name. Defaults to the first sheet."},
                "max_rows": {"type": "integer", "description": "Maximum rows to read (default 200).", "default": 200},
                "header_row": {"type": "boolean", "description": "Treat first row as headers. Default true.", "default": True},
            },
            "required": ["path"],
        },
    },
    {
        "name": "excel_list_sheets",
        "description": "List all sheet names in an .xlsx file.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string"},
            },
            "required": ["path"],
        },
    },
    {
        "name": "excel_write_sheet",
        "description": (
            "Write rows of data to a sheet in an .xlsx file. "
            "Creates the file/sheet if it doesn't exist. Optionally styles the header row."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Output .xlsx path."},
                "sheet": {"type": "string", "description": "Sheet name (default: Sheet1)."},
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column headers for the first row.",
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "List of rows. Each row is a list of cell values.",
                },
                "overwrite": {
                    "type": "boolean",
                    "description": "If true, replace the sheet if it exists. Default false (append).",
                    "default": False,
                },
            },
            "required": ["path", "rows"],
        },
    },
    {
        "name": "excel_update_cell",
        "description": "Update a single cell value in an .xlsx file.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "sheet": {"type": "string", "description": "Sheet name. Defaults to first sheet."},
                "cell": {"type": "string", "description": "Cell reference, e.g. 'B3' or 'A1'."},
                "value": {"description": "New cell value (string, number, or boolean)."},
            },
            "required": ["path", "cell", "value"],
        },
    },
    {
        "name": "excel_get_summary",
        "description": (
            "Get a numeric summary (min, max, sum, average, count) for a column in an .xlsx sheet."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "sheet": {"type": "string"},
                "column": {"type": "string", "description": "Column letter (e.g. 'B') or header name."},
                "skip_header": {"type": "boolean", "default": True},
            },
            "required": ["path", "column"],
        },
    },
    {
        "name": "excel_find_rows",
        "description": "Find rows where a column contains a specific value.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "sheet": {"type": "string"},
                "column": {"type": "string", "description": "Column letter (e.g. 'A') or header name."},
                "value": {"type": "string", "description": "Value to search for (case-insensitive substring match)."},
                "max_results": {"type": "integer", "default": 50},
            },
            "required": ["path", "column", "value"],
        },
    },
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _open_wb(path: str, read_only: bool = False) -> tuple:
    """Returns (workbook, is_new)."""
    p = Path(path)
    if p.exists():
        return load_workbook(path, read_only=read_only, data_only=True), False
    wb = Workbook()
    return wb, True


def _resolve_sheet(wb, sheet: str | None):
    if sheet and sheet in wb.sheetnames:
        return wb[sheet]
    return wb.active


def _col_index_from_header(ws, header_name: str) -> int | None:
    """Return 1-based column index for a header name, or None."""
    for cell in ws[1]:
        if str(cell.value).strip().lower() == header_name.strip().lower():
            return cell.column
    return None


def _resolve_col_index(ws, column: str) -> int:
    """Accept column letter ('B') or header name ('Revenue'). Returns 1-based int."""
    if len(column) <= 2 and column.isalpha():
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(column.upper())
    idx = _col_index_from_header(ws, column)
    if idx is None:
        raise ValueError(f"Column '{column}' not found in sheet headers.")
    return idx


# ---------------------------------------------------------------------------
# Implementations
# ---------------------------------------------------------------------------

def excel_read_sheet(
    path: str, sheet: str | None = None,
    max_rows: int = 200, header_row: bool = True,
) -> dict[str, Any]:
    wb, _ = _open_wb(path, read_only=True)
    ws = _resolve_sheet(wb, sheet)
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return {"path": path, "sheet": ws.title, "headers": [], "rows": [], "total_rows": 0}

    headers = [str(c) if c is not None else "" for c in all_rows[0]] if header_row else []
    data_rows = all_rows[1:max_rows + 1] if header_row else all_rows[:max_rows]
    rows = [[str(c) if c is not None else "" for c in r] for r in data_rows]

    return {
        "path": path,
        "sheet": ws.title,
        "headers": headers,
        "rows": rows,
        "total_rows": len(rows),
    }


def excel_list_sheets(path: str) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return {"path": path, "sheets": sheets, "count": len(sheets)}


def excel_write_sheet(
    path: str, rows: list[list], sheet: str = "Sheet1",
    headers: list[str] | None = None, overwrite: bool = False,
) -> dict[str, Any]:
    p = Path(path)
    if p.exists():
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            if overwrite:
                del wb[sheet]
                ws = wb.create_sheet(sheet)
            else:
                ws = wb[sheet]
        else:
            ws = wb.create_sheet(sheet)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        p.parent.mkdir(parents=True, exist_ok=True)

    start_row = ws.max_row + 1 if ws.max_row and not overwrite else 1

    if headers and (overwrite or ws.max_row == 1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1565C0")
            cell.font = Font(bold=True, color="FFFFFF")
        start_row += 1

    for r_idx, row in enumerate(rows, start_row):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    wb.save(path)
    return {"path": path, "sheet": sheet, "rows_written": len(rows), "status": "saved"}


def excel_update_cell(path: str, cell: str, value: Any, sheet: str | None = None) -> dict[str, Any]:
    wb = load_workbook(path)
    ws = _resolve_sheet(wb, sheet)
    ws[cell.upper()] = value
    wb.save(path)
    return {"path": path, "sheet": ws.title, "cell": cell.upper(), "value": value, "status": "updated"}


def excel_get_summary(
    path: str, column: str, sheet: str | None = None, skip_header: bool = True,
) -> dict[str, Any]:
    wb, _ = _open_wb(path, read_only=True)
    ws = _resolve_sheet(wb, sheet)
    col_idx = _resolve_col_index(ws, column)

    values = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if skip_header and r_idx == 1:
            continue
        val = row[col_idx - 1]
        if isinstance(val, (int, float)):
            values.append(val)
    wb.close()

    if not values:
        return {"column": column, "count": 0, "message": "No numeric values found."}

    return {
        "column": column,
        "count": len(values),
        "sum": round(sum(values), 4),
        "min": min(values),
        "max": max(values),
        "average": round(sum(values) / len(values), 4),
    }


def excel_find_rows(
    path: str, column: str, value: str,
    sheet: str | None = None, max_results: int = 50,
) -> dict[str, Any]:
    wb, _ = _open_wb(path, read_only=True)
    ws = _resolve_sheet(wb, sheet)
    col_idx = _resolve_col_index(ws, column)

    headers: list[str] = []
    matches: list[dict] = []

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if r_idx == 1:
            headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(row, 1)]
            continue
        cell_val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
        if value.lower() in cell_val.lower():
            matches.append({"row": r_idx, "data": dict(zip(headers, [str(c) if c else "" for c in row]))})
            if len(matches) >= max_results:
                break
    wb.close()

    return {"column": column, "search_value": value, "matches": matches, "total": len(matches)}


def dispatch(tool_name: str, tool_input: dict) -> dict[str, Any]:
    match tool_name:
        case "excel_read_sheet":  return excel_read_sheet(**tool_input)
        case "excel_list_sheets": return excel_list_sheets(**tool_input)
        case "excel_write_sheet": return excel_write_sheet(**tool_input)
        case "excel_update_cell": return excel_update_cell(**tool_input)
        case "excel_get_summary": return excel_get_summary(**tool_input)
        case "excel_find_rows":   return excel_find_rows(**tool_input)
        case _: raise ValueError(f"Unknown tool: {tool_name}")
